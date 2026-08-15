#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ZG-CAP-02 · gera a planilha de cotação (.xlsx) e a página de orçamento (.html)
a partir da mesma geometria usada no projeto executivo.

    python3 gerar_orcamento.py

Todo quantitativo sai da geometria. Os preços são FAIXA DE REFERÊNCIA de ordem
de grandeza, para triagem, e precisam ser substituídos pela cotação real.
"""
import json, math, os, datetime

AQUI = os.path.dirname(os.path.abspath(__file__))

# ----------------------------------------------------------------- geometria
D, H, FOLGA = 5000.0, 2292.0, 150.0        # mm, Bubble real do CAD
E1, E2 = 45.0, 89.0                         # faixa coberta, graus
W = 4800.0                                  # largura útil do tecido
NR, NP = 5, 5                               # arcos e longarinas
ROD, RWT = 50.0, 2.5                        # tubo dos arcos
POD, PWT = 32.0, 2.0                        # tubo das longarinas
TH_END = 6.0
BARRA = 6000.0                              # barra comercial

baseR = D / 2
zc = (H * H - baseR * baseR) / (2 * H)
R = H - zc
rc = R + FOLGA
W2 = W / 2
th0 = math.degrees(math.acos(min(W2 / rc, 1.0)))
de = E2 - E1
dRad = math.radians(de)
PISO = -480.0                               # piso do entorno vs. deck
sLim = (PISO + 100 - zc) / rc
aMax = (180 - math.degrees(math.asin(max(-1, min(1, sLim))))) - E2
aWork = aMax - 4
pedX = rc + 250

def massa_linear(od, wt, dens=2.70):
    """kg/m de tubo redondo. dens em g/cm³."""
    area = math.pi / 4 * (od ** 2 - (od - 2 * wt) ** 2)   # mm²
    return area * dens * 1e-3

KG_ARCO = massa_linear(ROD, RWT)
KG_LONG = massa_linear(POD, PWT)

L_ARCO = rc * math.radians(180 - 2 * TH_END)                    # mm, arco inteiro
L_ARCO_PANO = rc * math.radians(180 - 2 * th0)                  # trecho sob o pano
LONGARINAS = []
for j in range(NP):
    x = -W2 + 2 * W2 * j / (NP - 1)
    rho = math.sqrt(max(rc * rc - x * x, 1))
    LONGARINAS.append((x, rho, rho * dRad))
L_LONG_TOT = sum(l for _, _, l in LONGARINAS)

AREA_PANO = 2 * rc * W2 * dRad / 1e6                            # m²
PERIM_PANO = 2 * L_ARCO_PANO + 2 * LONGARINAS[0][2]             # mm
FLECHA = rc * (1 - math.cos(math.radians(de / (NR - 1)) / 2))

# barras de 6 m: cada arco consome 1 barra inteira + o resto
barras_arco = math.ceil(NR * L_ARCO * 1.06 / BARRA)
barras_long = math.ceil(L_LONG_TOT * 1.10 / BARRA)
kg_arco = barras_arco * BARRA / 1000 * KG_ARCO
kg_long = barras_long * BARRA / 1000 * KG_LONG
area_pintura = (math.pi * ROD / 1000 * NR * L_ARCO / 1000 +
                math.pi * POD / 1000 * L_LONG_TOT / 1000)
kg_cubo = 2 * 7.5
kg_eixo = 2 * (math.pi / 4 * 30 ** 2) * (pedX - rc + 140) * 7.9e-6
kg_ped = 2 * 0.30 * ((4 * 120 - 4 * 4.75) * 4.75 * 7.85e-3)
kg_base = 2 * (0.30 * 0.30 * 0.012 * 7850)

# ------------------------------------------------------- itens e faixas R$
# (grupo, item, descrição, unidade, quantidade, mín, provável, máx)
ITENS = [
 ("A. Estrutura de alumínio", "Tubo Ø50 × 2,5 · arcos", "6063-T6, barra de 6 m", "kg", kg_arco, 58, 72, 95),
 ("A. Estrutura de alumínio", "Tubo Ø32 × 2,0 · longarinas", "6063-T6, barra de 6 m", "kg", kg_long, 58, 72, 95),
 ("A. Estrutura de alumínio", "Perfil canaleta keder", "alumínio, ranhura Ø 6", "m", PERIM_PANO / 1000 * 1.08, 38, 58, 85),
 ("A. Estrutura de alumínio", "Luva de emenda interna", "alumínio maciço Ø 44", "un", NR, 25, 45, 80),
 ("B. Eixo, cubos e mancais", "Chapa inox 304 8 mm · cubo-aranha", "Ø 700, 5 braços", "kg", kg_cubo, 62, 82, 110),
 ("B. Eixo, cubos e mancais", "Barra inox 304 Ø30 h9 · eixo", "com rasgo de chaveta", "kg", kg_eixo, 58, 78, 105),
 ("B. Eixo, cubos e mancais", "Mancal de pé UCP206 inox", "Ø 30, autocompensador", "un", 2, 190, 340, 620),
 ("B. Eixo, cubos e mancais", "Parafusaria inox A4", "M6/M8/M10 + isolantes", "kit", 1, 260, 480, 850),
 ("C. Pedestais e fundação", "Tubo aço galv. 120×120×4,75", "berço do mancal", "kg", kg_ped, 14, 18, 25),
 ("C. Pedestais e fundação", "Chapa aço galv. 300×300×12", "placa de base", "kg", kg_base, 13, 17, 23),
 ("C. Pedestais e fundação", "Chumbador químico M12×160", "inox A4", "un", 8, 28, 45, 75),
 ("C. Pedestais e fundação", "Concreto do bloco 600×600×700", "2 blocos, usinado", "m³", 0.51, 550, 700, 900),
 ("D. Acionamento", "Atuador linear 24 V", "1.500 N, curso 350, IP66", "un", 2, 750, 1400, 2600),
 ("D. Acionamento", "Mola a gás inox", "550 N, compensação de peso", "un", 2, 180, 320, 550),
 ("D. Acionamento", "Central + controle remoto", "24 V, 2 canais, fim de curso", "un", 1, 380, 750, 1400),
 ("D. Acionamento", "Anemômetro", "recolhimento automático 45 km/h", "un", 1, 550, 1100, 2200),
 ("D. Acionamento", "Batentes e pinos de trava", "4 batentes + 2 pinos", "kit", 1, 250, 480, 900),
 ("D. Acionamento", "Cabeamento e quadro IP66", "24 V, disjuntor, fonte", "kit", 1, 300, 550, 950),
 ("E. Tecido e confecção", "Lona acrílica outdoor", "UV 95%, 600 g/m², areia", "m²", AREA_PANO * 1.10, 120, 210, 380),
 ("E. Tecido e confecção", "Cordão keder Ø6", "PVC, costurado na bainha", "m", PERIM_PANO / 1000 * 1.08, 6, 11, 20),
 ("E. Tecido e confecção", "Confecção e costura", "linha PTFE, solda de bordas", "m²", AREA_PANO, 65, 110, 180),
 ("E. Tecido e confecção", "Esticador de rosca inox M8", "curso 40 mm", "un", 10, 22, 40, 70),
 ("F. Serviços", "Calandragem dos arcos", "R 2.659, gabarito de piso", "peça", NR, 45, 90, 180),
 ("F. Serviços", "Calandragem das longarinas", "raio variável por estação", "peça", NP, 45, 90, 180),
 ("F. Serviços", "Corte a laser inox", "cubo-aranha", "peça", 2, 90, 180, 350),
 ("F. Serviços", "Usinagem do eixo e chaveta", "torno", "peça", 2, 120, 230, 400),
 ("F. Serviços", "Solda TIG alumínio e montagem", "bancada", "h", 40, 95, 140, 210),
 ("F. Serviços", "Pintura eletrostática", "dourado areia, pré-tratamento", "m²", area_pintura, 60, 95, 150),
 ("G. Campo e projeto", "Instalação em campo", "2 pessoas, 2 dias", "verba", 1, 2200, 4000, 7500),
 ("G. Campo e projeto", "Projeto estrutural e ART", "engenheiro responsável", "verba", 1, 1800, 3200, 6000),
]

FERROS = ("A. Estrutura de alumínio", "B. Eixo, cubos e mancais", "C. Pedestais e fundação")

def tot(sel=None):
    f = [i for i in ITENS if sel is None or i[0] in sel]
    return tuple(sum(i[4] * i[k] for i in f) for k in (5, 6, 7))

T_FERROS, T_TUDO = tot(FERROS), tot()
HOJE = datetime.date.today().strftime("%d/%m/%Y")

# ------------------------------------------------------------------- xlsx
def gerar_xlsx():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); ws = wb.active; ws.title = "Cotação ZG-CAP-02"
    ouro = "C9A84C"; escuro = "1B1F26"
    thin = Side(style="thin", color="D6D3CC")
    bd = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = "ZION GLAMPING · CAPOTA CHAPÉU RETRÁTIL ZG-CAP-02"
    ws["A1"].font = Font(bold=True, size=14, color=escuro)
    ws["A2"] = f"Planilha de cotação · Bubble Ø {D:.0f} mm no deck, altura {H:.0f} mm · emitida em {HOJE}"
    ws["A2"].font = Font(size=10, italic=True, color="6B7078")
    ws["A3"] = ("Preencha apenas a coluna SUA COTAÇÃO. As faixas são referência de ordem de "
                "grandeza para triagem e NÃO substituem orçamento do fornecedor.")
    ws["A3"].font = Font(size=10, color="A0522D")

    cab = ["Grupo", "Item", "Descrição", "Un.", "Qtd.", "R$ mín", "R$ provável",
           "R$ máx", "SUA COTAÇÃO R$/un", "TOTAL R$", "Fornecedor", "Prazo"]
    ws.append([]); ws.append(cab)
    hr = ws.max_row
    for c in range(1, len(cab) + 1):
        cel = ws.cell(row=hr, column=c)
        cel.font = Font(bold=True, size=9, color="FFFFFF")
        cel.fill = PatternFill("solid", fgColor=escuro)
        cel.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cel.border = bd

    inicio = hr + 1
    for g, it, ds, un, q, a, b, c in ITENS:
        ws.append([g, it, ds, un, round(q, 2), a, b, c, None, None, None, None])
        r = ws.max_row
        ws.cell(row=r, column=10).value = f"=IF(I{r}=\"\",E{r}*G{r},E{r}*I{r})"
        for k in range(1, 13):
            ws.cell(row=r, column=k).border = bd
        for k in (6, 7, 8, 9, 10):
            ws.cell(row=r, column=k).number_format = 'R$ #,##0.00'
        ws.cell(row=r, column=5).number_format = '#,##0.00'
        ws.cell(row=r, column=9).fill = PatternFill("solid", fgColor="FFF6DC")
    fim = ws.max_row

    def linha_total(rot, grupos, cor):
        ws.append([])
        r = ws.max_row + 1
        ws.cell(row=r, column=2, value=rot).font = Font(bold=True, size=11)
        # faixas somam QUANTIDADE × PREÇO, não preço unitário
        if grupos:
            filtro = "+".join(f'($A${inicio}:$A${fim}="{g}")' for g in grupos)
            faixa = f'SUMPRODUCT(({filtro})*$E${inicio}:$E${fim}*{{col}}${inicio}:{{col}}${fim})'
            coluna_j = "+".join(
                f'SUMIF($A${inicio}:$A${fim},"{g}",$J${inicio}:$J${fim})' for g in grupos)
        else:
            faixa = f'SUMPRODUCT($E${inicio}:$E${fim},{{col}}${inicio}:{{col}}${fim})'
            coluna_j = f'SUM($J${inicio}:$J${fim})'
        for col, dst in (("F", 6), ("G", 7), ("H", 8), ("J", 10)):
            expr = coluna_j if col == "J" else faixa.replace("{col}", col)
            ws.cell(row=r, column=dst, value="=" + expr)
            ws.cell(row=r, column=dst).number_format = 'R$ #,##0.00'
            ws.cell(row=r, column=dst).font = Font(bold=True)
            ws.cell(row=r, column=dst).fill = PatternFill("solid", fgColor=cor)
        return r

    linha_total("SUBTOTAL · SÓ OS FERROS (A+B+C)", FERROS, "F3EBD0")
    linha_total("TOTAL GERAL DO KIT", None, "E8DCB8")

    larg = [26, 34, 34, 7, 10, 12, 13, 12, 17, 15, 22, 12]
    for i, w_ in enumerate(larg, 1):
        ws.column_dimensions[get_column_letter(i)].width = w_
    ws.freeze_panes = f"A{inicio}"

    # aba de quantitativo
    q = wb.create_sheet("Quantitativo")
    q.append(["Grandeza", "Valor", "Unidade"])
    q["A1"].font = q["B1"].font = q["C1"].font = Font(bold=True)
    for lin in [
        ("Diâmetro da bolha no deck", round(D), "mm"),
        ("Altura da bolha", round(H), "mm"),
        ("Raio da esfera (derivado)", round(R, 1), "mm"),
        ("Centro da esfera / eixo do pivô", round(zc, 1), "mm"),
        ("Raio da casca", round(rc, 1), "mm"),
        ("Folga radial constante", round(FOLGA), "mm"),
        ("Vão entre pivôs", round(2 * rc), "mm"),
        ("Faixa coberta", f"{E1:.0f}° a {E2:.0f}°", "graus"),
        ("Curso homologado", round(aWork, 1), "graus"),
        ("Comprimento de cada arco", round(L_ARCO), "mm"),
        ("Comprimento total de longarina", round(L_LONG_TOT), "mm"),
        ("Barras de 6 m · arcos", barras_arco, "un"),
        ("Barras de 6 m · longarinas", barras_long, "un"),
        ("Massa de alumínio", round(kg_arco + kg_long, 1), "kg"),
        ("Massa de inox", round(kg_cubo + kg_eixo, 1), "kg"),
        ("Massa de aço galvanizado", round(kg_ped + kg_base, 1), "kg"),
        ("Área de tecido", round(AREA_PANO, 2), "m²"),
        ("Perímetro do pano", round(PERIM_PANO), "mm"),
        ("Área de pintura", round(area_pintura, 2), "m²"),
        ("Flecha do pano entre arcos", round(FLECHA), "mm"),
    ]:
        q.append(list(lin))
    for i, w_ in enumerate([40, 16, 12], 1):
        q.column_dimensions[get_column_letter(i)].width = w_

    caminho = os.path.join(AQUI, "ZG-CAP-02-cotacao.xlsx")
    wb.save(caminho)
    return caminho


if __name__ == "__main__":
    dados = {
        "itens": ITENS, "ferros": list(FERROS), "hoje": HOJE,
        "geo": {"D": D, "H": H, "R": R, "zc": zc, "rc": rc, "aWork": aWork,
                "areaPano": AREA_PANO, "kgAl": kg_arco + kg_long,
                "kgInox": kg_cubo + kg_eixo, "kgAco": kg_ped + kg_base,
                "barrasArco": barras_arco, "barrasLong": barras_long,
                "lArco": L_ARCO, "lLong": L_LONG_TOT, "areaPintura": area_pintura},
        "totais": {"ferros": T_FERROS, "tudo": T_TUDO},
    }
    with open(os.path.join(AQUI, "orcamento-dados.json"), "w", encoding="utf-8") as fh:
        json.dump(dados, fh, ensure_ascii=False, indent=1)
    print("xlsx:", gerar_xlsx())
    print(f"ferros  R$ {T_FERROS[0]:,.0f} / {T_FERROS[1]:,.0f} / {T_FERROS[2]:,.0f}")
    print(f"kit     R$ {T_TUDO[0]:,.0f} / {T_TUDO[1]:,.0f} / {T_TUDO[2]:,.0f}")
    print(f"alumínio {kg_arco + kg_long:.1f} kg · inox {kg_cubo + kg_eixo:.1f} kg · "
          f"aço {kg_ped + kg_base:.1f} kg · pano {AREA_PANO:.2f} m²")

    # injeta os dados no template e escreve a página final
    tpl = os.path.join(AQUI, "orcamento-template.html")
    if os.path.exists(tpl):
        with open(tpl, encoding="utf-8") as fh:
            html = fh.read()
        ini, fim_ = html.index("/*DADOS*/"), html.index("/*FIM*/") + len("/*FIM*/")
        html = html[:ini] + json.dumps(dados, ensure_ascii=False) + html[fim_:]
        saida = os.path.join(AQUI, "orcamento.html")
        with open(saida, "w", encoding="utf-8") as fh:
            fh.write(html)
        print("html:", saida)
